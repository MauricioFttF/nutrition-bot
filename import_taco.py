"""
import_taco.py

Script de execução ÚNICA — roda uma vez pra popular a aba "Alimentos"
da planilha com os dados da TACO 4ª edição (Unicamp). Não faz parte
do bot em execução contínua.

Uso:
    python import_taco.py

Requer as mesmas variáveis de ambiente do bot.py: SPREADSHEET_ID e
GOOGLE_CREDENTIALS_PATH (ou o caminho padrão credentials/service_account.json).

Estrutura esperada do arquivo (TACO 4ª edição, aba "CMVCol taco3"):
  - Cada linha de alimento tem um número inteiro na coluna A
    (Número do Alimento). Linhas de categoria/cabeçalho repetido/nota
    de rodapé NÃO têm número ali — é assim que distinguimos dado de
    não-dado, sem depender de contar linhas manualmente.
  - Colunas usadas: B=nome, D=kcal, F=proteína, G=lipídeos (gordura),
    I=carboidrato, J=fibra alimentar — todas já por 100g.
  - Valores ausentes aparecem como texto "NA", "Tr" (traço) ou "*" em
    vez de número — tratados como 0.
"""

from __future__ import annotations

import os

import openpyxl
from dotenv import load_dotenv

import sheets_client

load_dotenv()

_ARQUIVO_TACO = "data/taco.xlsx"
_ABA_ORIGEM = "CMVCol taco3"

_COL_NOME = 2
_COL_KCAL = 4
_COL_PROTEINA = 6
_COL_GORDURA = 7
_COL_CARBO = 9
_COL_FIBRA = 10


def _numero(valor) -> float:
    """Converte pra float; valores não-numéricos da TACO ('NA', 'Tr', '*') viram 0."""
    if isinstance(valor, (int, float)):
        return float(valor)
    return 0.0


def extrair_alimentos_taco(caminho_arquivo: str) -> list[dict]:
    """
    Lê o xlsx da TACO e retorna lista de dicts prontos pro formato da
    aba "Alimentos" (nome, fonte, porcao_base_g, kcal, proteina, carbo,
    gordura, fibra).
    """
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[_ABA_ORIGEM]

    alimentos = []
    for linha in range(1, ws.max_row + 1):
        numero_alimento = ws.cell(row=linha, column=1).value
        if not isinstance(numero_alimento, (int, float)):
            continue  # linha de categoria, cabeçalho repetido ou nota de rodapé

        nome = ws.cell(row=linha, column=_COL_NOME).value
        if not nome:
            continue

        alimentos.append({
            "nome": str(nome).strip(),
            "fonte": "TACO",
            "porcao_base_g": 100,
            "kcal": _numero(ws.cell(row=linha, column=_COL_KCAL).value),
            "proteina": _numero(ws.cell(row=linha, column=_COL_PROTEINA).value),
            "carbo": _numero(ws.cell(row=linha, column=_COL_CARBO).value),
            "gordura": _numero(ws.cell(row=linha, column=_COL_GORDURA).value),
            "fibra": _numero(ws.cell(row=linha, column=_COL_FIBRA).value),
        })

    return alimentos


def main() -> None:
    spreadsheet_id = os.environ["SPREADSHEET_ID"]
    credentials_path = os.environ.get(
        "GOOGLE_CREDENTIALS_PATH", "credentials/service_account.json"
    )

    print(f"Lendo {_ARQUIVO_TACO}...")
    alimentos = extrair_alimentos_taco(_ARQUIVO_TACO)
    print(f"{len(alimentos)} alimentos extraídos.")

    client = sheets_client.get_client(credentials_path)
    print("Enviando pra planilha (pode levar alguns segundos)...")
    sheets_client.importar_alimentos_em_lote(client, spreadsheet_id, alimentos)
    print("Import concluído.")


if __name__ == "__main__":
    main()